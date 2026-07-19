#!/usr/bin/env python3
"""
SEE Proof + Sign-off (Phase 4) - proofing we own, not the vendor's.

From a client's artwork + the booth spec, builds a branded PROOF that follows
SEE's unified proof standard (per Production's proof-standardization memo).

Two modes:
  * SINGLE ITEM (one artwork file): a one-page proof - artwork + a structured
    SPEC block + the automated preflight results (with a status legend) + a
    prominent disclaimer banner + a 3-option client sign-off + a consistent
    footer (prepped/QC/job#/version/fulfillment/page).
  * WHOLE JOB (several artwork files, or a folder, or --book): ONE multi-page
    PDF = a COVER/SUMMARY page (logo + job info INCLUDING the prepped/QC/job#/
    version/fulfillment section — shown ONCE here, not repeated per page — plus
    a table of every item + review instructions) followed by one page per item,
    each with just a Page X of Y footer.

Every proof is logged to proof_log.xlsx. Approve a single item to stamp + lock
the record; it refuses to approve a FAIL, or anything still carrying a
placeholder/blank value (no "TBD" or "Name here" reaches a client).

Usage:
    # one item
    python3 make_proof.py <artwork> [--spec booth_spec.json] [--panel NAME]
        [--job "Name"] [--prepped-by "Name"] [--qc-by "Name"]
        [--version V] [--fulfillment delivery|pickup] [--approve "Client Name"]
        [--ack-review "reason"]   # required to approve a NEEDS-REVIEW proof; recorded
                                  # (ignored — not stamped or logged — on any other verdict)
    # whole job (assembled document)
    python3 make_proof.py <art1> <art2> ...           # or a folder, or --book
        [--spec ...] [--prepped-by N] [--qc-by N] [--version V] [--fulfillment ...]
        [--allow-skips]           # accept (and disclose) files that could not be included

Exit codes: 0 = success; 1 = refusal (approval gate) or skipped files without
--allow-skips; 2 = usage error / unreadable input / no panel match.
"""
import sys, os, re, json, csv, glob, base64, subprocess, datetime, html, tempfile
import argparse
import errno
import fcntl, time
import proofer
import branding
try:
    import openpyxl
except Exception:
    openpyxl = None

RED = branding.RED
LOG = "proof_log.xlsx"                    # basename - resolve via default_log_path()
LOG_ENV = "SEE_PROOF_LOG"                 # overrides the log location
FALLBACK_CSV = "proof_log_fallback.csv"   # written when the xlsx can't be
# The one shared log-row contract (make_proof writes it, dashboard reads it).
LOG_HEADER = ["Date", "Job", "Job #", "Panel / Item", "File", "Verdict", "Status",
              "Proof version", "Prepped by", "QC'd by", "Approved by"]


def default_log_path():
    """The one place the proof log lives: $SEE_PROOF_LOG if set, else
    proof_log.xlsx at the REPO ROOT (one level above tools/). A cwd-relative
    path used to scatter records across job folders where the dashboard never
    found them."""
    env = os.environ.get(LOG_ENV)
    if env:
        return env
    return os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), LOG)


# errnos that mean "another process holds the flock" and are worth retrying:
# EWOULDBLOCK/EAGAIN per flock(2), plus EACCES which some platforms use for
# the same condition. Anything else (e.g. ENOTSUP/ENOLCK on network mounts
# that don't support flock at all) will NEVER clear, so retrying it just
# stalls every log write for the full timeout and then refuses the approval.
_LOCK_CONTENTION_ERRNOS = frozenset(
    e for e in (getattr(errno, n, None) for n in ("EWOULDBLOCK", "EAGAIN", "EACCES"))
    if e is not None)


class _FileLock:
    """Exclusive advisory lock (flock on LOG + '.lock') guarding the log's
    load->append->save cycle - unlocked concurrent runs dropped each other's
    rows, including APPROVED records the dashboard depends on. Blocks briefly,
    then raises TimeoutError rather than hanging a proof run forever. If the
    filesystem does not support flock (e.g. SEE_PROOF_LOG on a network mount
    raising ENOTSUP), proceeds WITHOUT the lock rather than failing - a
    single-writer setup still gets its row recorded."""

    def __init__(self, target, timeout=10.0):
        self.path = target + ".lock"
        self.timeout = timeout
        self.f = None
        self.locked = False

    def __enter__(self):
        self.f = open(self.path, "a", encoding="utf-8")
        deadline = time.time() + self.timeout
        while True:
            try:
                fcntl.flock(self.f, fcntl.LOCK_EX | fcntl.LOCK_NB)
                self.locked = True
                return self
            except OSError as e:
                if e.errno not in _LOCK_CONTENTION_ERRNOS:
                    # flock unsupported/broken here - not contention; go on
                    # unlocked instead of stalling the timeout and refusing
                    self.f.close()
                    self.f = None
                    return self
                if time.time() >= deadline:
                    self.f.close()
                    self.f = None
                    raise TimeoutError(f"could not lock {self.path} within {self.timeout:g}s "
                                       f"(another proof run holds it)")
                time.sleep(0.05)

    def __exit__(self, *exc):
        if self.f is None:
            return
        try:
            if self.locked:
                fcntl.flock(self.f, fcntl.LOCK_UN)
        finally:
            self.f.close()
# verdict badge colors come from the one check-status palette in proofer.BADGE
# (P3-3: the values were hand-duplicated here and could drift)
VCOL = {k: proofer.BADGE[k] for k in ("PASS", "REVIEW", "FAIL")}
VLABEL = {"PASS": "PASS", "REVIEW": "NEEDS REVIEW", "FAIL": "FAIL"}
CONTACT = branding.CONTACT
DISCLAIMER = ("This proof is for verifying CONTENT, LAYOUT, COLOR BREAK and SIZE only. On-screen color is "
              "not an exact match to the final printed piece. Check spelling, dimensions and finish "
              "carefully — your approval releases this file to print.")
ART_EXT = (".pdf", ".ai", ".eps") + proofer.RASTER_EXT

# literal placeholder / unfilled markers that must never reach a client
# (deliberately NOT 'n/a' - that is a legitimate value)
PLACEHOLDER_RE = re.compile(
    r"\b(tbd|tba|todo|name here|placeholder|lorem|xxx+|fpo|tk|fill ?in|change ?me|"
    r"client name|your (?:name|logo|text))\b|[<>]|\?\?\?", re.I)


def looks_placeholder(v):
    """True if a value carries literal placeholder/markup text (the memo's
    'no placeholder reaches the client' failure, e.g. 'TBD', 'Name here')."""
    return bool(v) and bool(PLACEHOLDER_RE.search(str(v)))


def is_blank(v):
    """True if a value is effectively empty (None / '' / a bare dash)."""
    return v is None or str(v).strip() in ("", "—", "-")


def panel_specs(panel, spec, version=None):
    """Ordered (label, value) rows for the client-facing spec block - the fields
    SEE's proof standard requires, pulled from the single-source booth spec.
    Optional fields (finishing_type / quantity / seams / tracking_id / rev) are
    read from the panel when present, else a sensible default or '—'. Pure."""
    w, h = panel.get("w"), panel.get("h")
    finish_size = f'{h:g}" H × {w:g}" W' if (w and h) else "—"
    sided = str(panel.get("sided", "")).strip().lower()
    sided_label = {"single": "Single-sided", "double": "Double-sided"}.get(sided, panel.get("sided") or "—")
    seams = panel.get("seams")
    rev = panel.get("rev") or version or spec.get("job", {}).get("version") or "—"
    return [
        ("Item / tracking #", panel.get("tracking_id") or panel.get("name") or "—"),
        ("Finish size (H × W)", finish_size),
        ("Material", panel.get("finish") or "—"),
        ("Finishing type", panel.get("finishing_type") or "—"),
        ("Quantity", str(panel.get("quantity", 1))),
        ("Sides", sided_label),
        ("Seams", "—" if seams in (None, "") else str(seams)),
        ("Revision", str(rev)),
    ]


def proof_readiness(specs, prepped_by, qc_by, material):
    """Client-readiness gate (the memo's 'no placeholder reaches the client' rule
    + 'QC/Prepped by = an actual name, never a placeholder'). Returns
    (placeholders, missing):
      placeholders - values carrying literal placeholder text; must NEVER ship.
      missing      - required fields (Prepped by, QC'd by, Material) not yet set.
    Blank OPTIONAL spec fields (finishing type / seams) are not 'missing' here -
    the caller surfaces those as a soft note. Pure function."""
    placeholders, missing = [], []
    for label, val in specs:
        if looks_placeholder(val):
            placeholders.append(f"{label} = '{val}'")
    for label, val in (("Prepped by", prepped_by), ("QC'd by", qc_by)):
        if is_blank(val):
            missing.append(label)
        elif looks_placeholder(val):
            placeholders.append(f"{label} = '{val}'")
    if is_blank(material):
        missing.append("Material")
    return placeholders, missing


def job_readiness(spec, job=None, version=None):
    """Placeholder scan over the JOB-level fields that render on client proofs
    (job name, client, show, job #, version) - these used to bypass the
    per-panel placeholder gate. Returns the same \"Label = 'value'\" strings
    proof_readiness produces, so callers can merge the two lists. Pure."""
    j = spec.get("job", {})
    fields = [("Job name", job if job is not None else j.get("name")),
              ("Client", j.get("client")),
              ("Show", j.get("show")),
              ("Job #", j.get("job_number")),
              ("Proof version", version if version is not None else j.get("version"))]
    return [f"{label} = '{val}'" for label, val in fields if looks_placeholder(val)]


def job_totals(items):
    """(# graphics, # pieces) for the cover - graphics = number of items,
    pieces = sum of each item's quantity (default 1). Pure function."""
    pieces = 0
    for it in items:
        q = it["panel"].get("quantity", 1)
        try:
            pieces += int(q)
        except (TypeError, ValueError):
            pieces += 1
    return len(items), pieces


def cover_rows(items):
    """Per-item rows for the cover summary table: (item, size, material, sides,
    qty). Pure function."""
    rows = []
    for it in items:
        p = it["panel"]
        w, h = p.get("w"), p.get("h")
        size = f'{h:g}" × {w:g}"' if (w and h) else "—"
        sided = str(p.get("sided", "")).strip().lower()
        sides = {"single": "1", "double": "2"}.get(sided, str(p.get("sided") or "—"))
        rows.append((p.get("tracking_id") or p.get("name") or "—", size,
                     p.get("finish") or "—", sides, str(p.get("quantity", 1))))
    return rows


def thumbnail(path, ext, tag=""):
    """Rasterize the artwork to a UNIQUE temp PNG and return its path, or None.

    The output must be verifiably produced by THIS run (gs returncode 0 and a
    non-empty file that gs itself created) - a fixed cwd name used to let a
    stale PNG from a crashed run, or a concurrent run in the same directory,
    become another job's artwork on a client sign-off proof. None means the
    proof shows '(preview unavailable)' instead of a wrong image. The caller
    removes the returned file when done."""
    fd, out = tempfile.mkstemp(prefix=f"_proof_thumb{tag}_", suffix=".png")
    os.close(fd)
    try:
        if ext in proofer.RASTER_EXT:
            from PIL import Image
            im = Image.open(path)
            if im.mode == "CMYK":
                im = im.convert("RGB")
            im.thumbnail((1000, 1000))
            im.save(out)
        else:
            os.remove(out)  # gs must CREATE the file - never trust a pre-existing one
            p = subprocess.run(["gs", "-q", "-sDEVICE=png16m", "-r60", "-dFirstPage=1",
                                "-dLastPage=1", "-o", out, path], capture_output=True)
            if p.returncode != 0:
                raise RuntimeError(f"gs exited {p.returncode}")
        if not (os.path.exists(out) and os.path.getsize(out) > 0):
            raise RuntimeError("no thumbnail output produced")
        return out
    except Exception:
        _cleanup(out)
        return None


def b64img(p):
    return "data:image/png;base64," + base64.b64encode(open(p, "rb").read()).decode() if p else ""


# P3-3: one logo lookup for every generated document — branding.logo_data_uri
# (which searches only the canonical assets/see_logo.png locations, never a
# cwd wildcard that could embed a client's own logo file).
_logo_data_uri = branding.logo_data_uri


def log_proof(job, job_no, panel, fname, verdict, status, version, prepped, qc, approver,
              today=None):
    """Append one row to the proof log; returns where the row landed.
    `today` (a datetime.date, injectable for tests — mirrors dashboard.py)
    defaults to the real date.

    The whole load->append->save cycle holds an exclusive flock, and NO record
    is ever lost: when openpyxl is missing or the workbook can't be loaded or
    saved (locked/corrupt), the same row is appended to proof_log_fallback.csv
    next to the log and the returned note names it. Raises only when NEITHER
    destination could be written - the approve path then refuses to stamp
    (via _log_proof_safe)."""
    path = default_log_path()
    today = today or datetime.date.today()
    row = [today.isoformat() if hasattr(today, "isoformat") else str(today),
           job, job_no or "", panel,
           os.path.basename(fname), verdict, status, version or "",
           prepped or "", qc or "", approver or ""]
    with _FileLock(path):
        if openpyxl:
            try:
                if os.path.exists(path):
                    wb = openpyxl.load_workbook(path); ws = wb.active
                else:
                    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Proofs"
                    ws.append(LOG_HEADER)
                ws.append(row)
                wb.save(path)
                return path
            except Exception as e:
                xlsx_err = f"{type(e).__name__}: {e}"
        else:
            xlsx_err = "openpyxl missing"
        # CSV fallback - the record must not be lost
        csv_path = os.path.join(os.path.dirname(path) or ".", FALLBACK_CSV)
        new = not os.path.exists(csv_path)
        with open(csv_path, "a", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            if new:
                w.writerow(LOG_HEADER)
            w.writerow(row)
        return f"{csv_path} (xlsx unavailable: {xlsx_err})"


def _annotate_last_log_row(fname, panel, status, note):
    """Append `note` to the Status of the MOST RECENT xlsx log row matching
    (File, Panel / Item, Status) — the approve path logs BEFORE rendering (an
    approval that can't be logged must not stamp), so when the PDF render then
    fails its row needs a follow-up annotation (P1-4). Same flock as
    log_proof. Returns True when the row was updated in place; False when it
    couldn't be (openpyxl missing, the row landed in the CSV fallback, the
    workbook is locked/corrupt) so the caller appends a follow-up row
    instead — the outcome is recorded either way."""
    if not openpyxl:
        return False
    path = default_log_path()
    base = os.path.basename(fname)
    f_col = LOG_HEADER.index("File") + 1
    p_col = LOG_HEADER.index("Panel / Item") + 1
    s_col = LOG_HEADER.index("Status") + 1
    try:
        with _FileLock(path):
            if not os.path.exists(path):
                return False
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            for r in range(ws.max_row, 1, -1):
                if (ws.cell(r, f_col).value == base and ws.cell(r, p_col).value == panel
                        and ws.cell(r, s_col).value == status):
                    ws.cell(r, s_col).value = status + note
                    wb.save(path)
                    return True
            return False
    except Exception:
        return False


CSS_PROOF = f"""
  @page {{ size: letter portrait; margin: 0.5in; }}
  body {{ font-family: {branding.FONT_STACK}; color:#1a1a1a; font-size:12px; margin:0; }}
  .page {{ page-break-after: always; }}
  .page:last-child {{ page-break-after: auto; }}
  .pill {{ background:{RED}; color:#fff; display:inline-block; padding:5px 14px; border-radius:16px; font-weight:700; font-size:11px; }}
  h1 {{ font-size:19px; margin:6px 0 1px; }}
  .meta {{ color:#555; font-size:11px; margin:1px 0 0; }}
  .verdict {{ display:inline-block; color:#fff; padding:4px 12px; border-radius:7px; font-weight:700; font-size:13px; }}
  .legend {{ font-size:10px; color:#666; margin-left:8px; }}
  .legend .dot {{ display:inline-block; width:9px; height:9px; border-radius:50%; margin:0 3px 0 9px; vertical-align:baseline; }}
  .banner {{ margin:6px 0; padding:6px 12px; background:#FFF4E5; border:1px solid #F7941E; border-left:6px solid #F7941E;
            border-radius:5px; color:#7a4a00; font-size:11px; font-weight:600; line-height:1.3; }}
  .caution {{ margin:6px 0; padding:6px 12px; background:#fde8e8; border:1px solid {RED}; border-left:6px solid {RED};
             border-radius:5px; color:#7a0d12; font-size:11px; font-weight:700; line-height:1.3; }}
  .cols {{ display:flex; gap:15px; margin-top:4px; }}
  .art {{ flex:0 0 40%; border:1px solid #ddd; border-radius:6px; padding:6px; text-align:center; background:#fafafa; align-self:flex-start; }}
  .art img {{ max-width:100%; max-height:300px; }}
  .noimg {{ color:#999; padding:40px 0; }}
  .right {{ flex:1; }}
  .blk {{ font-size:10px; text-transform:uppercase; letter-spacing:.04em; color:#888; font-weight:700; margin:0 0 4px; }}
  table {{ width:100%; border-collapse:collapse; }}
  table.spec {{ margin-bottom:8px; }}
  table.spec td {{ padding:3px 8px; border-bottom:1px solid #eee; font-size:11px; vertical-align:top; }}
  td.sl {{ color:#666; width:42%; }}
  td.sv {{ font-weight:700; }}
  table.chk th {{ background:#f3f3f3; text-align:left; padding:5px 8px; border-bottom:2px solid #ccc; font-size:9.5px; text-transform:uppercase; }}
  table.chk td {{ padding:3px 8px; border-bottom:1px solid #ececec; vertical-align:top; }}
  td.ck {{ font-weight:700; width:20%; }}
  .b {{ color:#fff; padding:2px 9px; border-radius:10px; font-weight:700; font-size:9.5px; }}
  .msg {{ font-size:10px; }}
  ol.fixlist {{ margin:7px 0 0; padding-left:18px; }}
  ol.fixlist li {{ font-size:10px; margin:2px 0; color:#7a4a00; }}
  .signbox {{ margin-top:9px; border:1.5px solid #bbb; border-radius:8px; padding:9px 13px; }}
  .sign .st {{ font-weight:700; color:{RED}; margin-bottom:5px; }}
  .sign .opt {{ margin-bottom:3px; font-size:12px; }}
  .sign .lines {{ margin:7px 0 6px; }}
  .sign .chg {{ color:#555; }}
  .stamp {{ display:inline-block; border:3px solid #2E9E40; color:#2E9E40; font-weight:800; font-size:17px; padding:7px 16px; border-radius:8px; letter-spacing:.04em; }}
  .locknote {{ color:#7a0d12; font-size:11px; margin-top:8px; }}
  footer {{ margin-top:9px; border-top:1px solid #ddd; padding-top:5px; }}
  .ftgrid {{ display:flex; flex-wrap:wrap; gap:6px 18px; font-size:10px; }}
  .ftgrid div span {{ display:block; text-transform:uppercase; letter-spacing:.03em; color:#999; font-size:8.5px; font-weight:700; }}
  .ftgrid div b {{ font-size:11px; }}
  .contact {{ color:#999; font-size:9px; margin-top:7px; }}
  /* cover */
  .brandrow {{ display:flex; justify-content:space-between; align-items:center; }}
  .wordmark {{ font-size:18px; font-weight:800; color:{RED}; letter-spacing:.01em; }}
  .logo {{ height:56px; width:auto; display:block; margin-bottom:4px; }}
  .logosm {{ height:30px; width:auto; }}
  .phead {{ display:flex; justify-content:space-between; align-items:center; }}
  .coverhead {{ color:#999; font-size:9px; margin-top:2px; }}
  h1.cv {{ font-size:23px; margin:16px 0 2px; }}
  .jobgrid {{ display:flex; flex-wrap:wrap; gap:7px 26px; margin:10px 0 4px; }}
  .jobgrid div span {{ display:block; text-transform:uppercase; letter-spacing:.03em; color:#999; font-size:8.5px; font-weight:700; }}
  .jobgrid div b {{ font-size:12.5px; }}
  .totals {{ margin:14px 0 6px; font-size:13px; }}
  .totals b {{ color:{RED}; }}
  table.summary th {{ background:{RED}; color:#fff; text-align:left; padding:7px 9px; font-size:10px; text-transform:uppercase; }}
  table.summary td {{ padding:6px 9px; border-bottom:1px solid #eaeaea; font-size:11px; }}
  table.summary tr:nth-child(even) td {{ background:#fafafa; }}
  table.summary .muted {{ color:#c0392b; font-weight:700; }}
  .howto {{ margin-top:15px; border:1px solid #ddd; border-radius:7px; padding:11px 14px; background:#f7f9fb; font-size:11px; line-height:1.5; }}
  .howto b {{ color:{RED}; }}
"""

HEAD = '<!doctype html><html><head><meta charset="utf-8"><style>' + CSS_PROOF + '</style></head><body>'
FOOT = '</body></html>'


def _item_footer(meta, today, job_no, full=True):
    """Item-page footer. `full` renders the whole prepped/QC/job#/version/
    fulfillment grid (the standalone single-item proof, where this is the ONE
    place that section appears). In the whole-job document that section lives
    once on the cover, so item pages pass full=False and carry only the page
    number + contact line — no repeated metadata block on every page."""
    page, pages = meta.get("page", 1), meta.get("pages", 1)
    if not full:
        return f"""<footer>
        <div class="ftgrid"><div><span>Page</span><b>{page} of {pages}</b></div></div>
        <div class="contact">{CONTACT}</div>
      </footer>"""
    version = meta.get("version") or "—"
    prepped = meta.get("prepped_by") or "—"
    qc = meta.get("qc_by") or "—"
    fulfillment = (meta.get("fulfillment") or "").title() or "—"
    return f"""<footer>
        <div class="ftgrid">
          <div><span>Prepped by</span><b>{html.escape(str(prepped))}</b></div>
          <div><span>QC'd by</span><b>{html.escape(str(qc))}</b></div>
          <div><span>Date issued</span><b>{today}</b></div>
          <div><span>Job #</span><b>{html.escape(str(job_no))}</b></div>
          <div><span>Proof version</span><b>{html.escape(str(version))}</b></div>
          <div><span>Fulfillment</span><b>{html.escape(str(fulfillment))}</b></div>
          <div><span>Page</span><b>{page} of {pages}</b></div>
        </div>
        <div class="contact">{CONTACT}</div>
      </footer>"""


def _item_body(job, res, spec, thumb_b64, approve, meta, logo="", today=None, full_footer=True):
    """One item's page (no <html>/<body> wrapper) - a <section class='page'>.
    `logo` (a data URI) shows a small mark in the header for a standalone proof;
    in the job document the cover carries the logo, so item pages pass ''.
    `today` (a preformatted display string, injectable for golden tests —
    mirrors dashboard.py's injectable today) defaults to the real date.
    `full_footer` False (whole-job document) prints just the page number in the
    footer — the prepped/QC/job#/version/fulfillment section rides the cover once."""
    today = today or datetime.date.today().strftime("%B %d, %Y")
    verdict = res["verdict"]
    panel = res["panel"]
    specs = meta["specs"]
    job_no = spec.get("job", {}).get("job_number") or spec.get("job", {}).get("estimate") or "—"
    version = meta.get("version") or "—"

    spec_rows = "".join(
        f'<tr><td class="sl">{html.escape(l)}</td><td class="sv">{html.escape(str(v))}</td></tr>'
        for l, v in specs)
    chk_rows = ""
    for k in proofer.ORDER:
        if k in res["results"]:
            st, msg = res["results"][k]
            chk_rows += (f'<tr><td class="ck">{k.title()}</td>'
                         f'<td><span class="b" style="background:{proofer.BADGE[st]}">{st}</span></td>'
                         f'<td class="msg">{html.escape(msg)}</td></tr>')
    fixes = res.get("fixes") or []
    fix_block = ""
    if fixes:
        flis = "".join(f'<li><b>{html.escape(f["check"].title())}:</b> {html.escape(f["text"])}</li>'
                       for f in fixes)
        fix_block = f'<div class="blk">What to change</div><ol class="fixlist">{flis}</ol>'
    if approve:
        ack = meta.get("ack_review")
        ack_html = (f'<div class="locknote">NEEDS-REVIEW items acknowledged before approval — '
                    f'reason: {html.escape(str(ack))}</div>' if ack else '')
        signoff = (f'<div class="stamp">APPROVED &nbsp;·&nbsp; {html.escape(approve)} &nbsp;·&nbsp; {today}</div>'
                   f'{ack_html}'
                   f'<div class="locknote">Locked on approval. Any change after this requires written approval '
                   f'and may trigger an add-on charge.</div>')
    else:
        signoff = (
            '<div class="sign"><div class="st">Client approval</div>'
            '<div class="opt">&#9744;&nbsp; Approved as is</div>'
            '<div class="opt">&#9744;&nbsp; Approved with changes noted below</div>'
            '<div class="opt">&#9744;&nbsp; Revisions required — please resubmit for approval</div>'
            '<div class="lines">Signature ____________________________ &nbsp; Printed name ____________________ &nbsp; Date __________</div>'
            '<div class="chg">Notes / changes: _______________________________________________________________________</div></div>')
    img = f'<img src="{thumb_b64}">' if thumb_b64 else '<div class="noimg">(preview unavailable)</div>'
    caution = ""
    if not approve and (meta.get("placeholders") or meta.get("missing")):
        items = "; ".join(meta["placeholders"] + [f"{m} not set" for m in meta["missing"]])
        caution = (f'<div class="caution">&#9888; DRAFT — not client-ready: {html.escape(items)}. '
                   f'Resolve before sending to the client.</div>')

    logo_html = f'<img class="logosm" src="{logo}">' if logo else ''
    return f"""<section class="page">
      <div class="phead"><div class="pill">Artwork Proof — for client approval</div>{logo_html}</div>
      <h1>{html.escape(job)} &nbsp;—&nbsp; Item {html.escape(panel['name'])}</h1>
      <div class="meta">Proof version {html.escape(str(version))} &nbsp;·&nbsp; {today} &nbsp;·&nbsp; checked against the booth spec &nbsp;·&nbsp;
        <span class="verdict" style="background:{VCOL[verdict]}">{VLABEL[verdict]}</span>
        <span class="legend">
          <span class="dot" style="background:#2E9E40"></span>Pass
          <span class="dot" style="background:#F7941E"></span>Needs review
          <span class="dot" style="background:#E31D3D"></span>Fail</span>
      </div>
      <div class="banner">{DISCLAIMER}</div>
      {caution}
      <div class="cols">
        <div class="art">{img}</div>
        <div class="right">
          <div class="blk">Specifications</div>
          <table class="spec">{spec_rows}</table>
          <div class="blk">Automated preflight checks</div>
          <table class="chk"><thead><tr><th>Check</th><th>Result</th><th>Detail</th></tr></thead><tbody>{chk_rows}</tbody></table>
          {fix_block}
        </div>
      </div>
      <div class="signbox">{signoff}</div>
      {_item_footer(meta, today, job_no, full=full_footer)}
    </section>"""


def build_proof_html(job, res, spec, thumb_b64, approve, meta, today=None):
    """Single-item proof (full HTML document). `today` is injectable (P2-6)."""
    return HEAD + _item_body(job, res, spec, thumb_b64, approve, meta,
                             logo=_logo_data_uri(), today=today) + FOOT


def _cover_body(job, spec, items, meta, unmatched=None, today=None):
    """The job COVER / summary page (a <section class='page'>). `unmatched`
    (list of 'filename (reason)' strings) renders a red caution block - a
    client signing off the job must SEE which files the proof does not cover,
    not just a console line the designer saw. `today` (display string) is
    injectable for golden tests; defaults to the real date."""
    today = today or datetime.date.today().strftime("%B %d, %Y")
    logo = _logo_data_uri()
    brand = (f'<img class="logo" src="{logo}">' if logo
             else '<div class="wordmark">Southeast Exhibits &amp; Events</div>')
    j = spec.get("job", {})
    job_no = j.get("job_number") or j.get("estimate") or "—"
    version = meta.get("version") or j.get("version") or "—"
    due = j.get("due_date")
    due_txt = "" if (is_blank(due) or looks_placeholder(due)) else f" Please return the signed proof by <b>{html.escape(str(due))}</b>."
    n_graphics, n_pieces = job_totals(items)

    fields = [("Client", j.get("client")), ("Show", j.get("show")), ("Booth", j.get("booth_size")),
              ("Job #", job_no), ("Proof version", version), ("Date issued", today),
              ("Prepped by", meta.get("prepped_by")), ("QC'd by", meta.get("qc_by")),
              ("Fulfillment", (meta.get("fulfillment") or "").title())]
    jobgrid = "".join(f'<div><span>{html.escape(l)}</span><b>{html.escape(str(v or "—"))}</b></div>'
                      for l, v in fields)

    srows = ""
    for i, (name, size, material, sides, qty) in enumerate(cover_rows(items), 1):
        mcls = ' class="muted"' if looks_placeholder(material) else ""
        srows += (f'<tr><td>{i}</td><td><b>{html.escape(name)}</b></td><td>{html.escape(size)}</td>'
                  f'<td{mcls}>{html.escape(material)}</td><td>{html.escape(sides)}</td><td>{html.escape(qty)}</td></tr>')

    skipped_html = ""
    if unmatched:
        names = "; ".join(html.escape(str(u)) for u in unmatched)
        skipped_html = (f'<div class="caution">&#9888; NOT INCLUDED in this proof: {names} — '
                        f'these files could not be checked. Do not sign off this job until '
                        f'every graphic is accounted for.</div>')

    return f"""<section class="page cover">
      <div class="brandrow">
        <div>{brand}
          <div class="coverhead">Orlando | Las Vegas | Atlanta | NJ/NY | Dallas &nbsp;·&nbsp; SouthEastExhibit.com</div></div>
        <div class="pill">Client Proof</div>
      </div>
      <h1 class="cv">{html.escape(job)}</h1>
      <div class="jobgrid">{jobgrid}</div>
      <div class="totals">This proof covers <b>{n_graphics}</b> graphic(s) — <b>{n_pieces}</b> piece(s) total.</div>
      {skipped_html}
      <table class="summary"><thead><tr><th>#</th><th>Item</th><th>Finish size (H × W)</th><th>Material</th><th>Sides</th><th>Qty</th></tr></thead>
        <tbody>{srows}</tbody></table>
      <div class="howto"><b>How to review:</b> Each graphic is on its own page that follows. For every item,
        check the artwork, spelling, dimensions and finish, then mark one box — <b>Approved as is</b>,
        <b>Approved with changes</b>, or <b>Revisions required</b> — and sign and date it.{due_txt}
        Your approval releases that file to print.</div>
      <footer><div class="contact">{CONTACT}</div>
        <div class="ftgrid" style="margin-top:5px"><div><span>Page</span><b>1 of {meta.get('pages', 1)}</b></div></div>
      </footer>
    </section>"""


def build_job_html(job, spec, items, approve, base_meta, unmatched=None, today=None):
    """Whole-job document: cover page + one page per item, with Page X of Y.
    `unmatched` files are disclosed in a caution block on the cover.
    `today` is injectable (P2-6) and threads to the cover and every item page."""
    pages = len(items) + 1
    base_meta = dict(base_meta, pages=pages)
    out = HEAD + _cover_body(job, spec, items, base_meta, unmatched, today=today)
    for idx, it in enumerate(items):
        meta = dict(base_meta, specs=it["specs"], placeholders=it["placeholders"],
                    missing=it["missing"], page=idx + 2, pages=pages)
        out += _item_body(job, it["res"], spec, it["thumb_b64"], approve, meta,
                          today=today, full_footer=False)
    return out + FOOT


# ---------- CLI ----------
def collect_files(raw):
    """Expand any directory arguments into the artwork files they contain."""
    files = []
    for a in raw:
        if os.path.isdir(a):
            for f in sorted(glob.glob(os.path.join(a, "*"))):
                if os.path.splitext(f)[1].lower() in ART_EXT:
                    files.append(f)
        else:
            files.append(a)
    return files


def main(argv=None):
    # argparse (the old hand-rolled loop turned a typo'd flag into a "file",
    # silently flipping to job mode and DROPPING --approve; a trailing flag
    # raised IndexError). Unknown flags now exit 2 with usage.
    ap = argparse.ArgumentParser(
        prog="make_proof.py", allow_abbrev=False,
        description="Build a branded client proof: one page per artwork item, "
                    "checked against the booth spec. Several files (or a folder, "
                    "or --book) build the whole-job document.")
    ap.add_argument("artwork", nargs="+", help="artwork file(s) and/or folder(s)")
    ap.add_argument("--spec", help="booth spec JSON (default: auto-discovered)")
    ap.add_argument("--panel", help="panel name (single item; job mode matches by filename)")
    ap.add_argument("--job", help="job name (default: from the spec)")
    ap.add_argument("--approve", metavar="NAME",
                    help="stamp APPROVED by NAME (single item; gated + logged)")
    ap.add_argument("--ack-review", dest="ack_review", metavar="REASON",
                    help="acknowledge a NEEDS-REVIEW verdict with a recorded reason")
    ap.add_argument("--prepped-by", "--prepped", dest="prepped_by", metavar="NAME")
    ap.add_argument("--qc-by", "--qc", dest="qc_by", metavar="NAME")
    ap.add_argument("--version", dest="version", metavar="V",
                    help="proof version (also suffixes the output name)")
    ap.add_argument("--fulfillment", choices=("delivery", "pickup"))
    ap.add_argument("--book", action="store_true",
                    help="build the whole-job document even for a single file")
    ap.add_argument("--allow-skips", dest="allow_skips", action="store_true",
                    help="job mode: exit 0 even when files were skipped (still disclosed)")
    a = ap.parse_args(argv)

    files = collect_files(a.artwork)
    if not files:
        ap.error("no artwork files found in the given path(s)")
    spec_path = a.spec or proofer.find_default_spec()
    print(f"Spec: {spec_path}")
    try:
        with open(spec_path, encoding="utf-8") as f:
            spec = json.load(f)
    except (OSError, ValueError) as e:
        print(f"could not read the booth spec {spec_path}: {e}", file=sys.stderr)
        sys.exit(2)
    job = a.job or spec.get("job", {}).get("name", "Untitled job")
    version = a.version or spec.get("job", {}).get("version")
    job_no = spec.get("job", {}).get("job_number") or spec.get("job", {}).get("estimate")
    base_meta = {"prepped_by": a.prepped_by, "qc_by": a.qc_by, "version": version,
                 "fulfillment": a.fulfillment, "ack_review": a.ack_review}

    job_mode = len(files) > 1 or a.book
    if job_mode and a.panel and len(files) != 1:
        print("--panel names ONE panel, but job mode matches each file to a panel by "
              "FILENAME. Run one file at a time with --panel, or drop --panel.",
              file=sys.stderr)
        sys.exit(2)
    if job_mode:
        rc = build_job_proof(files, spec, job, job_no, a.approve, base_meta, a.panel,
                             allow_skips=a.allow_skips)
    else:
        rc = build_single_proof(files[0], spec, job, job_no, a.approve, base_meta, a.panel)
    if rc:
        sys.exit(rc)


def safe_stem(name):
    """Filename stem sanitized for output names. Keeps '-' and '.' so distinct
    artworks stay distinct ('F 1.pdf' -> F_1, 'F-1.pdf' -> F-1 - the old
    collapse-everything-to-_ made both write the SAME proof file)."""
    return re.sub(r"[^A-Za-z0-9._-]+", "_", str(name)).strip("_") or "artwork"


def _backup_existing(path):
    """Rename an existing output aside with a timestamp instead of overwriting
    it (an APPROVED proof is a signed record - never silently replaced).
    Returns the backup path or None."""
    if not os.path.exists(path):
        return None
    root, ext = os.path.splitext(path)
    ts = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
    bak = f"{root}_superseded_{ts}{ext}"
    os.replace(path, bak)
    print(f"note: existing {os.path.basename(path)} renamed to {os.path.basename(bak)} "
          f"(an approved proof is never silently overwritten)")
    return bak


def build_single_proof(fname, spec, job, job_no, approve, base_meta, panel_arg):
    """Returns an exit status: 0 success, 1 approval refusal, 2 unreadable/no
    panel match (the refusal/error paths used to 'return' -> exit 0, so a
    scripted approve-then-email flow treated a refusal as success)."""
    ext = os.path.splitext(fname)[1].lower()
    try:
        res = proofer.run_checks(fname, spec, panel_arg)
    except Exception as e:
        print("could not read file:", e); return 2
    if res and res.get("error"):
        print(res["error"]); return 2
    if not res:
        print("could not match to a panel — re-run with --panel NAME"); return 2
    panel = res["panel"]
    refusal, specs, placeholders, missing = approval_decision(
        res, spec, job, fname, approve, base_meta)
    if refusal:
        print(refusal); return 1
    ack_review = base_meta.get("ack_review")
    if ack_review and res["verdict"] != "REVIEW":
        # P0-8: an acknowledgment only applies to a NEEDS-REVIEW verdict. On
        # any other verdict it is ignored ENTIRELY (never stamped on the proof
        # or logged) — otherwise `--approve X --ack-review "TBD"` on a clean
        # PASS would print a false "NEEDS-REVIEW items acknowledged" line on
        # the client-facing proof and record it in the log. When the reason
        # WILL be recorded (REVIEW verdict + --approve), approval_decision has
        # already refused a blank/placeholder reason above.
        print(f"note: --ack-review ignored — preflight verdict is {res['verdict']}, "
              f"nothing was under review (the reason is not stamped on the proof or logged).")
        ack_review = None
        base_meta = dict(base_meta, ack_review=None)   # keep it off the rendered proof too

    status = "APPROVED" if approve else f"PROOFED ({res['verdict']})"
    if approve and ack_review:
        status = f"APPROVED (REVIEW acknowledged: {ack_review})"
    if approve:
        # log BEFORE stamping: an approval that cannot be logged must not ship
        logged, log_ok = _log_proof_safe(job, job_no, panel["name"], fname, res["verdict"],
                                         status, base_meta.get("version"),
                                         base_meta.get("prepped_by"), base_meta.get("qc_by"),
                                         approve)
        if not log_ok:
            print(f"⛔ Refusing to stamp APPROVED: the approval could not be logged {logged}.\n"
                  f"   Every approval must be recorded in {default_log_path()} (or its CSV "
                  f"fallback) — fix the log location/permissions and re-run.")
            return 1

    thumb = thumbnail(fname, ext)
    try:
        meta = dict(base_meta, specs=specs, placeholders=placeholders, missing=missing, page=1, pages=1)
        page = build_proof_html(job, res, spec, b64img(thumb), approve, meta)
    finally:
        _cleanup(thumb)

    base = safe_stem(os.path.splitext(os.path.basename(fname))[0])
    ver = base_meta.get("version")
    vtag = f"_v{safe_stem(ver)}" if ver else ""
    suffix = f"_PROOF{vtag}" + ("_APPROVED" if approve else "")
    hp = os.path.abspath(base + suffix + ".html")
    pp = os.path.abspath(base + suffix + ".pdf")
    if approve:
        # a previously APPROVED proof at this name is a signed record - keep it
        _backup_existing(hp)
        _backup_existing(pp)
    with open(hp, "w", encoding="utf-8") as f:
        f.write(page)
    ok = proofer.render_pdf(hp, pp)
    if approve and not ok:
        # P1-4: the approve path logged BEFORE the render (see above), so a
        # failed render would leave its row promising a PDF that was never
        # produced. Annotate that row's Status in place (same wording as the
        # non-approve path below); when it can't be updated (CSV fallback,
        # locked workbook) append a follow-up row instead.
        note = " — PDF render failed (HTML only)"
        if _annotate_last_log_row(fname, panel["name"], status, note):
            logged = f"{logged} (Status annotated: PDF render failed, HTML only)"
        else:
            logged, _ = _log_proof_safe(job, job_no, panel["name"], fname, res["verdict"],
                                        status + note, base_meta.get("version"),
                                        base_meta.get("prepped_by"), base_meta.get("qc_by"),
                                        approve)
    if not approve:
        # log AFTER the render so the row's Status reflects what actually exists
        # (the approve path logs BEFORE stamping instead - see above)
        if not ok:
            status += " — PDF render failed (HTML only)"
        logged, _ = _log_proof_safe(job, job_no, panel["name"], fname, res["verdict"], status,
                                    base_meta.get("version"), base_meta.get("prepped_by"),
                                    base_meta.get("qc_by"), approve)
    print(f"\nItem {panel['name']}  ·  verdict {res['verdict']}  ·  " +
          (f"APPROVED by {approve}" if approve else "awaiting client sign-off"))
    if not approve and (placeholders or missing):
        print("⚠  NOT client-ready yet — resolve before sending:\n   - "
              + "\n   - ".join(placeholders + [f"{m} not set" for m in missing]))
    print("Proof sheet:", os.path.basename(pp) if ok else os.path.basename(hp) + " (open + print to PDF)")
    print("Logged to  :", logged)
    return 0


def build_job_proof(files, spec, job, job_no, approve, base_meta, panel_arg, allow_skips=False):
    """Returns an exit status: 0 success, 1 skipped files without --allow-skips,
    2 nothing matched at all. `panel_arg` is honored only when exactly one file
    was given (job mode otherwise matches by filename - main() rejects the
    ambiguous combination up front)."""
    if approve:
        print("note: --approve is for a single item; the job document is a draft for per-item sign-off. Ignoring --approve.")
        approve = None
    panel_index = {p["name"]: i for i, p in enumerate(spec.get("panels", []))}
    job_placeholders = job_readiness(spec, job, base_meta.get("version"))
    items, unmatched = [], []
    explicit_panel = panel_arg if len(files) == 1 else None
    for n, fname in enumerate(files):
        ext = os.path.splitext(fname)[1].lower()
        try:
            res = proofer.run_checks(fname, spec, explicit_panel)
        except Exception as e:
            unmatched.append(f"{os.path.basename(fname)} (could not be read: {e})"); continue
        if res and res.get("error"):
            unmatched.append(f"{os.path.basename(fname)} ({res['error']})"); continue
        if not res:
            unmatched.append(f"{os.path.basename(fname)} (no matching panel)"); continue
        panel = res["panel"]
        specs = panel_specs(panel, spec, base_meta.get("version"))
        placeholders, missing = proof_readiness(specs, base_meta.get("prepped_by"),
                                                base_meta.get("qc_by"), panel.get("finish"))
        placeholders = job_placeholders + placeholders
        if panel.get("needs_confirm"):
            missing = missing + ["panel dimensions UNVERIFIED (AI/OCR-sourced — confirm in the booth file)"]
        thumb = thumbnail(fname, ext, tag=str(n))
        items.append({"panel": panel, "res": res, "specs": specs, "fname": fname,
                      "placeholders": placeholders, "missing": missing,
                      "thumb_b64": b64img(thumb), "_thumb": thumb})
    if not items:
        print("no files matched a panel — name them after the panel (e.g. F1.pdf) or use the single-item mode with --panel")
        if unmatched:
            print("  skipped:", ", ".join(unmatched))
        return 2
    items.sort(key=lambda it: panel_index.get(it["panel"]["name"], 999))

    try:
        html_doc = build_job_html(job, spec, items, approve, base_meta, unmatched)
    finally:
        for it in items:
            _cleanup(it.get("_thumb"))

    base = safe_stem(job)
    ver = base_meta.get("version")
    vtag = f"_v{safe_stem(ver)}" if ver else ""
    hp = os.path.abspath(base + "_JOB_PROOF" + vtag + ".html")
    pp = os.path.abspath(base + "_JOB_PROOF" + vtag + ".pdf")
    with open(hp, "w", encoding="utf-8") as f:
        f.write(html_doc)
    ok = proofer.render_pdf(hp, pp)
    # log AFTER the render so the rows' Status reflects what actually exists
    status = "PROOFED (job doc)" + ("" if ok else " — PDF render failed (HTML only)")
    logged = None
    for it in items:
        logged, _ = _log_proof_safe(job, job_no, it["panel"]["name"], it["fname"],
                                    it["res"]["verdict"], status, base_meta.get("version"),
                                    base_meta.get("prepped_by"), base_meta.get("qc_by"), None)
    n_graphics, n_pieces = job_totals(items)
    print(f"\nJOB PROOF · {job}")
    print(f"  {n_graphics} item(s), {n_pieces} piece(s) · {len(items) + 1} pages (cover + {len(items)})")
    for it in items:
        flag = "  ⚠ not client-ready" if (it["placeholders"] or it["missing"]) else ""
        how = it["res"].get("how", "?")
        print(f"    - {it['panel']['name']:14} {it['res']['verdict']} [{how}]{flag}")
    if unmatched:
        print("  ⚠ NOT INCLUDED (disclosed on the proof cover):", ", ".join(unmatched))
    print("Document:", os.path.basename(pp) if ok else os.path.basename(hp) + " (open + print to PDF)")
    if logged:
        print("Logged to  :", logged)
    if unmatched and not allow_skips:
        print(f"{len(unmatched)} file(s) could not be included — fix them, or re-run with "
              f"--allow-skips to accept the disclosed omission. Exiting nonzero.")
        return 1
    return 0


def approval_decision(res, spec, job, fname, approve, base_meta):
    """The COMPLETE approval-gate decision for one checked item, as a PURE
    function — no I/O, no Chrome/gs — so invariant 4 (approval must refuse
    when a check fails or a measurement is unconfirmed) is testable end to
    end (P2-3). Combines the readiness scan (placeholder values on panel and
    job fields, missing prepped-by/QC names, needs_confirm dimensions) with
    the approver-name validation and the _approval_block refusals
    (FAIL / non-PASS size / un-acked REVIEW; ack_review comes from base_meta).

    Returns (refusal, specs, placeholders, missing): `refusal` is a printable
    refusal message, or None when approval may proceed (or when no approval
    was requested); specs/placeholders/missing feed the proof page and the
    not-client-ready warning either way."""
    panel = res["panel"]
    specs = panel_specs(panel, spec, base_meta.get("version"))
    placeholders, missing = proof_readiness(specs, base_meta.get("prepped_by"),
                                            base_meta.get("qc_by"), panel.get("finish"))
    placeholders = job_readiness(spec, job, base_meta.get("version")) + placeholders
    if panel.get("needs_confirm"):
        missing = missing + ["panel dimensions UNVERIFIED (AI/OCR-sourced — confirm in the booth file)"]
    refusal = None
    if approve is not None and (is_blank(approve) or looks_placeholder(approve)):
        refusal = (f'⛔ Refusing to stamp APPROVED: approver "{approve}" is blank or a placeholder — '
                   f"--approve needs the real client approver's name.")
    elif approve:
        refusal = _approval_block(res, placeholders, missing, os.path.basename(fname),
                                  base_meta.get("ack_review"))
    return refusal, specs, placeholders, missing


def _approval_block(res, placeholders, missing, fname, ack_review=None):
    """Return a refusal message if this item can't be locked-approved, else None.

    Invariant 4: approval must refuse when a check fails or a measurement is
    unconfirmed. Beyond the FAIL refusal: a size check that did not PASS is
    ALWAYS refused (no flag overrides an unverified/wrong finished size); any
    other NEEDS-REVIEW verdict needs an explicit --ack-review \"reason\",
    which is recorded on the proof and in the log row."""
    if res["verdict"] == "FAIL":
        return (f"⛔ Refusing to stamp APPROVED: {fname} FAILS preflight "
                f"({', '.join(k for k, v in res['results'].items() if v[0] == 'FAIL')}). Fix the FAIL(s) first.")
    size_st, size_msg = res["results"].get("size", ("NA", "size was not checked"))
    if size_st != "PASS":
        return (f"⛔ Refusing to stamp APPROVED: the finished size is unverified or wrong — "
                f"size check is {size_st}: {size_msg}\n   A measurement that did not PASS can "
                f"never be approved (--ack-review does not override size); fix the file or "
                f"verify the size first.")
    if placeholders:
        return ("⛔ Refusing to stamp APPROVED: placeholder/blank values would reach the client:\n   - "
                + "\n   - ".join(placeholders) + "\n   Fill them in the booth spec first.")
    if missing:
        return ("⛔ Refusing to stamp APPROVED: not client-ready — " + ", ".join(missing)
                + ".\n   Confirm these in the booth file (and provide --prepped-by / --qc-by) before approving.")
    if res["verdict"] == "REVIEW":
        warns = [k for k, v in res["results"].items() if v[0] == "WARN"]
        if is_blank(ack_review) or looks_placeholder(ack_review):
            return ("⛔ Refusing to stamp APPROVED: preflight verdict is NEEDS REVIEW — WARN on "
                    + ", ".join(warns) + ".\n   Review those items, then re-run with "
                    "--ack-review \"reason\" to approve with a recorded acknowledgment.")
    return None


def _log_proof_safe(*args):
    """(logged, ok) - ok is False only when the row could not be persisted
    ANYWHERE (the xlsx path AND the CSV fallback both failed). Used on the
    approval path: an approval that cannot be logged must not stamp
    (invariant 4 adjacent - the 'stamped and logged' promise)."""
    try:
        return log_proof(*args), True
    except Exception as e:
        return f"(log write failed: {e})", False


def _cleanup(path):
    if path:
        try:
            os.remove(path)
        except OSError:
            pass


if __name__ == "__main__":
    main()
