"""P1-4: proof log integrity — flock-guarded writes, CSV fallback, one fixed
location ($SEE_PROOF_LOG, else repo root), log-after-render for non-approve
proofs. Rows written by make_proof.log_proof are round-tripped through
dashboard.read_proof_log. No gs/Chrome/network needed (openpyxl only).
"""
import csv
import datetime
import fcntl
import os
import threading
import time

import pytest

import dashboard
import make_proof as mp


def _log_row(panel="F1", verdict="PASS", approver=None, status="PROOFED (PASS)"):
    return mp.log_proof("Booth Build", "1001", panel, "F1.pdf", verdict, status,
                        "C1", "A. Tech", "M. Palumbo", approver)


def _canned_res():
    return {"panel": {"name": "F1", "w": 78, "h": 134, "finish": "Fabric"},
            "how": "named explicitly", "info": {"kind": "pdf"},
            "results": {"size": ("PASS", "ok"), "color": ("PASS", "CMYK")},
            "verdict": "PASS", "fixes": []}


CLEAN_SPEC = {"job": {"name": "Booth Build", "client": "Acme Co", "show": "NAB",
                      "job_number": "1001", "version": "C1"},
              "settings": {},
              "panels": [{"name": "F1", "w": 78, "h": 134, "finish": "Fabric"}]}
META = {"prepped_by": "A. Tech", "qc_by": "M. Palumbo", "version": "C1",
        "fulfillment": "delivery", "ack_review": None}


# ---------- location ----------
def test_default_location_is_repo_root_env_var_overrides(monkeypatch):
    monkeypatch.delenv("SEE_PROOF_LOG", raising=False)
    repo_root = os.path.dirname(os.path.dirname(os.path.abspath(mp.__file__)))
    assert mp.default_log_path() == os.path.join(repo_root, "proof_log.xlsx")
    monkeypatch.setenv("SEE_PROOF_LOG", "/shared/drive/mylog.xlsx")
    assert mp.default_log_path() == "/shared/drive/mylog.xlsx"


# ---------- round trip ----------
def test_rows_round_trip_into_dashboard(tmp_path, monkeypatch):
    log = tmp_path / "proof_log.xlsx"
    monkeypatch.setenv("SEE_PROOF_LOG", str(log))
    assert _log_row(panel="F1") == str(log)
    _log_row(panel="F2", verdict="REVIEW", approver="Jane Client")
    rows = dashboard.read_proof_log(str(log))[0]["1001"]
    assert [r["Panel / Item"] for r in rows] == ["F1", "F2"]     # file order kept
    assert rows[0]["Date"] == datetime.date.today().isoformat()
    assert rows[1]["Approved by"] == "Jane Client"
    import openpyxl
    header = [c.value for c in next(openpyxl.load_workbook(str(log)).active.iter_rows())]
    assert header == mp.LOG_HEADER                               # shared contract


# ---------- locking ----------
def test_interleaved_writes_block_on_the_lock_and_both_persist(tmp_path, monkeypatch):
    log = tmp_path / "proof_log.xlsx"
    monkeypatch.setenv("SEE_PROOF_LOG", str(log))
    holder = open(str(log) + ".lock", "w")
    fcntl.flock(holder, fcntl.LOCK_EX)                  # simulate a concurrent run
    done = []
    t = threading.Thread(target=lambda: done.append(_log_row(panel="F1")))
    t.start()
    time.sleep(0.25)
    assert not done and not log.exists()                # blocked, not dropped
    fcntl.flock(holder, fcntl.LOCK_UN)
    holder.close()
    t.join(timeout=10)
    assert done == [str(log)]
    _log_row(panel="F2")
    rows = dashboard.read_proof_log(str(log))[0]["1001"]
    assert [r["Panel / Item"] for r in rows] == ["F1", "F2"]   # nobody's row lost


# ---------- CSV fallback ----------
def test_openpyxl_missing_row_lands_in_csv_and_note_names_it(tmp_path, monkeypatch):
    monkeypatch.setenv("SEE_PROOF_LOG", str(tmp_path / "proof_log.xlsx"))
    monkeypatch.setattr(mp, "openpyxl", None)
    logged = _log_row(approver="Jane Client")
    assert "proof_log_fallback.csv" in logged and "openpyxl missing" in logged
    rows = list(csv.reader(open(tmp_path / "proof_log_fallback.csv")))
    assert rows[0] == mp.LOG_HEADER
    assert rows[1][mp.LOG_HEADER.index("Panel / Item")] == "F1"
    assert rows[1][mp.LOG_HEADER.index("Approved by")] == "Jane Client"


class _LockedOpenpyxl:
    """Stands in for openpyxl whose workbook is locked/corrupt on load & save."""
    @staticmethod
    def load_workbook(*a, **k):
        raise OSError("xlsx is locked by Excel")

    @staticmethod
    def Workbook(*a, **k):
        raise OSError("xlsx is locked by Excel")


def test_workbook_save_failure_falls_back_to_csv_no_traceback(tmp_path, monkeypatch):
    monkeypatch.setenv("SEE_PROOF_LOG", str(tmp_path / "proof_log.xlsx"))
    monkeypatch.setattr(mp, "openpyxl", _LockedOpenpyxl)
    logged = _log_row()
    assert "proof_log_fallback.csv" in logged and "locked by Excel" in logged
    assert (tmp_path / "proof_log_fallback.csv").exists()


def test_approval_with_locked_xlsx_stamps_via_csv_and_still_renders(tmp_path, monkeypatch, capsys):
    monkeypatch.chdir(tmp_path)
    monkeypatch.setenv("SEE_PROOF_LOG", str(tmp_path / "proof_log.xlsx"))
    monkeypatch.setattr(mp, "openpyxl", _LockedOpenpyxl)
    monkeypatch.setattr(mp.proofer, "run_checks", lambda *a, **k: _canned_res())
    rendered = []
    monkeypatch.setattr(mp.proofer, "render_pdf",
                        lambda hp, pp: (rendered.append(hp), False)[1])
    mp.build_single_proof("F1.pdf", CLEAN_SPEC, "Booth Build", "1001",
                          "Jane Client", dict(META), None)      # no SystemExit
    assert rendered, "render must still happen after the fallback log"
    assert os.path.exists("F1_PROOF_APPROVED.html")
    csv_text = open("proof_log_fallback.csv").read()
    assert "Jane Client" in csv_text


# ---------- log-after-render (non-approve) ----------
def test_non_approve_logs_after_render_and_records_the_outcome(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    log = tmp_path / "proof_log.xlsx"
    monkeypatch.setenv("SEE_PROOF_LOG", str(log))
    monkeypatch.setattr(mp.proofer, "run_checks", lambda *a, **k: _canned_res())
    order = []
    monkeypatch.setattr(mp.proofer, "render_pdf",
                        lambda hp, pp: (order.append("render"), False)[1])
    real_log = mp.log_proof
    monkeypatch.setattr(mp, "log_proof",
                        lambda *a: (order.append("log"), real_log(*a))[1])
    mp.build_single_proof("F1.pdf", CLEAN_SPEC, "Booth Build", "1001",
                          None, dict(META), None)
    assert order == ["render", "log"]                   # the log can't claim a
    rows = dashboard.read_proof_log(str(log))[0]["1001"]   # proof that wasn't made
    assert "PDF render failed (HTML only)" in rows[-1]["Status"]


# ---------- dashboard discovery / merging ----------
def test_find_logs_env_var_and_jobs_dir_merge(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    sub = tmp_path / "jobs" / "jobA"
    sub.mkdir(parents=True)
    monkeypatch.setenv("SEE_PROOF_LOG", str(sub / "proof_log.xlsx"))
    _log_row()                                          # a log written inside a job folder
    # env now points elsewhere (nothing there): only the jobs-dir glob can find it
    monkeypatch.setenv("SEE_PROOF_LOG", str(tmp_path / "elsewhere.xlsx"))
    paths = dashboard.find_logs(jobs_dir=str(tmp_path / "jobs"))
    assert [os.path.realpath(p) for p in paths] == [os.path.realpath(str(sub / "proof_log.xlsx"))]
    merged = {}
    for p in paths:
        for k, v in dashboard.read_proof_log(p)[0].items():
            merged.setdefault(k, []).extend(v)
    assert [r["Panel / Item"] for r in merged["1001"]] == ["F1"]


def test_find_logs_explicit_log_flag_wins(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    log = tmp_path / "custom.xlsx"
    monkeypatch.setenv("SEE_PROOF_LOG", str(log))
    _log_row()
    monkeypatch.setenv("SEE_PROOF_LOG", str(tmp_path / "other.xlsx"))
    assert dashboard.find_logs(explicit=str(log)) == [str(log)]
    assert dashboard.find_logs(explicit=str(tmp_path / "missing.xlsx")) == []
