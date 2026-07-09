"""P0-8: the approval gate (invariant 4 - approval must refuse when a check
fails or a measurement is unconfirmed).

Covers _approval_block directly plus the build_single_proof CLI path with
proofer.run_checks monkeypatched to canned results, so no real PDF, gs,
Chrome, or network is needed (openpyxl is used for the log round-trip).
"""
import os

import pytest

import make_proof as mp


def canned_res(verdict="PASS", results=None, panel=None):
    results = results or {"size": ("PASS", "matches full + bleed"),
                          "color": ("PASS", "CMYK")}
    panel = panel or {"name": "F1", "w": 78, "h": 134, "finish": "Fabric"}
    return {"panel": panel, "how": "named explicitly", "info": {"kind": "pdf"},
            "results": results, "verdict": verdict, "fixes": []}


CLEAN_SPEC = {"job": {"name": "Booth Build", "client": "Acme Co", "show": "NAB",
                      "job_number": "1001", "version": "C1"},
              "settings": {},
              "panels": [{"name": "F1", "w": 78, "h": 134, "finish": "Fabric"}]}

META = {"prepped_by": "A. Tech", "qc_by": "M. Palumbo", "version": "C1",
        "fulfillment": "delivery", "ack_review": None}


# ---------- _approval_block (pure) ----------
def test_fail_verdict_refused():
    res = canned_res("FAIL", {"size": ("PASS", "ok"), "color": ("FAIL", "RGB")})
    msg = mp._approval_block(res, [], [], "f.pdf")
    assert msg and "FAILS preflight" in msg and "color" in msg


def test_size_warn_refused_names_unverified_measurement():
    res = canned_res("REVIEW", {"size": ("WARN", "cannot verify finished size"),
                                "color": ("PASS", "CMYK")})
    msg = mp._approval_block(res, [], [], "f.pdf")
    assert msg and "size" in msg and "unverified or wrong" in msg
    assert "cannot verify finished size" in msg


def test_size_warn_refused_even_with_ack_review():
    res = canned_res("REVIEW", {"size": ("WARN", "no bleed detected"),
                                "color": ("PASS", "CMYK")})
    msg = mp._approval_block(res, [], [], "f.pdf", ack_review="looked at it")
    assert msg and "--ack-review does not override size" in msg


def test_review_without_ack_refused_listing_warns_and_flag():
    res = canned_res("REVIEW", {"size": ("PASS", "ok"),
                                "spelling": ("WARN", "1 word to review"),
                                "fonts": ("WARN", "live fonts")})
    msg = mp._approval_block(res, [], [], "f.pdf")
    assert msg and "NEEDS REVIEW" in msg
    assert "spelling" in msg and "fonts" in msg     # lists the WARN checks
    assert "--ack-review" in msg                    # names the flag to add


def test_review_with_ack_reason_allows():
    res = canned_res("REVIEW", {"size": ("PASS", "ok"),
                                "spelling": ("WARN", "1 word to review")})
    assert mp._approval_block(res, [], [], "f.pdf", ack_review="checked manually") is None


def test_review_with_placeholder_ack_reason_still_refused():
    res = canned_res("REVIEW", {"size": ("PASS", "ok"), "spelling": ("WARN", "x")})
    assert mp._approval_block(res, [], [], "f.pdf", ack_review="TBD") is not None


def test_placeholders_and_missing_still_refused():
    res = canned_res()
    assert "placeholder" in mp._approval_block(res, ["Material = 'TBD'"], [], "f.pdf")
    assert "not client-ready" in mp._approval_block(res, [], ["Prepped by"], "f.pdf")


def test_pass_verdict_clean_is_approvable():
    assert mp._approval_block(canned_res(), [], [], "f.pdf") is None


# ---------- approval_decision (the pure, extracted gate - P2-3) ----------
def _decide(res, approve, ack=None, spec=None):
    meta = dict(META, ack_review=ack)
    return mp.approval_decision(res, spec or CLEAN_SPEC, "Booth Build",
                                "F1.pdf", approve, meta)


def test_decision_fail_verdict_refuses():
    res = canned_res("FAIL", {"size": ("PASS", "ok"), "color": ("FAIL", "RGB")})
    refusal, specs, placeholders, missing = _decide(res, "Jane Client")
    assert refusal and "FAILS preflight" in refusal and "color" in refusal
    assert specs and placeholders == [] and missing == []


def test_decision_size_warn_refuses_naming_measurement():
    res = canned_res("REVIEW", {"size": ("WARN", "cannot verify finished size")})
    refusal = _decide(res, "Jane Client")[0]
    assert refusal and "unverified or wrong" in refusal
    assert "cannot verify finished size" in refusal


def test_decision_review_needs_ack_then_allows_with_reason():
    res = canned_res("REVIEW", {"size": ("PASS", "ok"),
                                "spelling": ("WARN", "1 word to review")})
    refusal = _decide(res, "Jane Client")[0]
    assert refusal and "--ack-review" in refusal and "spelling" in refusal
    assert _decide(res, "Jane Client", ack="checked manually")[0] is None


def test_decision_placeholder_approver_refuses():
    refusal = _decide(canned_res(), "TBD")[0]
    assert refusal and "blank or a placeholder" in refusal
    assert _decide(canned_res(), "  ")[0] is not None


def test_decision_needs_confirm_panel_refuses_and_reports_missing():
    res = canned_res(panel={"name": "F1", "w": 78, "h": 134, "finish": "Fabric",
                            "needs_confirm": True})
    refusal, _, _, missing = _decide(res, "Jane Client")
    assert refusal and "UNVERIFIED" in refusal
    assert any("UNVERIFIED" in m for m in missing)


def test_decision_job_placeholder_refuses():
    spec = {"job": dict(CLEAN_SPEC["job"], client="TBD"), "settings": {},
            "panels": CLEAN_SPEC["panels"]}
    refusal, _, placeholders, _ = _decide(canned_res(), "Jane Client", spec=spec)
    assert refusal and "Client = 'TBD'" in refusal
    assert any("Client" in p for p in placeholders)


def test_decision_clean_pass_approves():
    refusal, specs, placeholders, missing = _decide(canned_res(), "Jane Client")
    assert refusal is None and placeholders == [] and missing == []


def test_decision_no_approve_never_refuses_but_still_reports_readiness():
    # without --approve nothing is gated, but the not-client-ready data still
    # feeds the proof page and console warning
    res = canned_res("FAIL", {"size": ("FAIL", "wrong size")},
                     panel={"name": "F1", "w": 78, "h": 134, "finish": "TBD"})
    refusal, _, placeholders, _ = _decide(res, None)
    assert refusal is None
    assert any("TBD" in p for p in placeholders)


# ---------- job-level placeholder gate ----------
def test_job_readiness_flags_job_fields():
    spec = {"job": {"name": "Booth", "client": "TBD", "show": "Your Name Here",
                    "job_number": "1001", "version": "C1"}}
    hits = mp.job_readiness(spec)
    assert any("Client" in h for h in hits)
    assert any("Show" in h for h in hits)
    assert mp.job_readiness(CLEAN_SPEC) == []


def test_placeholder_re_extensions():
    for v in ("FPO", "tk", "fill in", "fillin", "Change me", "Client Name", "your logo"):
        assert mp.looks_placeholder(v), v
    for v in ("n/a", "N/A", "Fabric", "Atkins Design", "milk"):
        assert not mp.looks_placeholder(v), v


# ---------- CLI-level: build_single_proof ----------
@pytest.fixture
def in_tmp(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    return tmp_path


def _run(monkeypatch, res, approve, ack=None, spec=None):
    monkeypatch.setattr(mp.proofer, "run_checks", lambda *a, **k: res)
    meta = dict(META, ack_review=ack)
    return mp.build_single_proof("F1.pdf", spec or CLEAN_SPEC, "Booth Build", "1001",
                                 approve, meta, None)


def test_approve_canned_fail_cli_refuses_no_artifacts_no_log(in_tmp, monkeypatch, capsys):
    # P2-3's CLI-level scenario: canned FAIL + render_pdf -> False. The
    # refusal must print, exit 1, and leave NO stamped proof and NO log row.
    monkeypatch.setattr(mp.proofer, "render_pdf", lambda *a, **k: False)
    res = canned_res("FAIL", {"size": ("PASS", "ok"), "color": ("FAIL", "RGB")})
    rc = _run(monkeypatch, res, "Jane Client")
    assert rc == 1
    assert "FAILS preflight" in capsys.readouterr().out
    assert not [f for f in os.listdir(".") if "APPROVED" in f]
    assert not os.path.exists("proof_log.xlsx")
    assert not os.path.exists("proof_log_fallback.csv")


def test_approve_placeholder_approver_refused(in_tmp, monkeypatch, capsys):
    rc = _run(monkeypatch, canned_res(), "TBD")
    assert rc == 1                                       # P1-7: refusal -> status 1
    assert "blank or a placeholder" in capsys.readouterr().out
    assert not os.path.exists("F1_PROOF_vC1_APPROVED.html")


def test_approve_review_without_ack_refused_cli(in_tmp, monkeypatch, capsys):
    res = canned_res("REVIEW", {"size": ("PASS", "ok"), "spelling": ("WARN", "x")})
    rc = _run(monkeypatch, res, "Jane Client")
    assert rc == 1
    assert "--ack-review" in capsys.readouterr().out
    assert not os.path.exists("F1_PROOF_vC1_APPROVED.html")
    assert not os.path.exists("proof_log.xlsx"), "a refused approval must not log"


def test_approve_needs_confirm_panel_refused(in_tmp, monkeypatch, capsys):
    res = canned_res(panel={"name": "F1", "w": 78, "h": 134, "finish": "Fabric",
                            "needs_confirm": True})
    rc = _run(monkeypatch, res, "Jane Client")
    assert rc == 1
    assert "UNVERIFIED" in capsys.readouterr().out


def test_approve_job_placeholder_refused(in_tmp, monkeypatch, capsys):
    spec = {"job": dict(CLEAN_SPEC["job"], client="TBD"), "settings": {},
            "panels": CLEAN_SPEC["panels"]}
    rc = _run(monkeypatch, canned_res(), "Jane Client", spec=spec)
    assert rc == 1
    assert "Client = 'TBD'" in capsys.readouterr().out


def test_approve_openpyxl_missing_lands_in_csv_fallback(in_tmp, monkeypatch, capsys):
    # P1-4: a missing openpyxl no longer blocks the approval - the row lands in
    # the CSV fallback instead (no record is ever lost), and the note names it
    monkeypatch.setattr(mp, "openpyxl", None)
    assert _run(monkeypatch, canned_res(), "Jane Client") == 0
    assert os.path.exists("F1_PROOF_vC1_APPROVED.html")
    csv_text = open("proof_log_fallback.csv", encoding="utf-8").read()
    assert "Jane Client" in csv_text and "F1" in csv_text
    assert "proof_log_fallback.csv" in capsys.readouterr().out


def test_approve_refused_when_no_log_destination_writable(in_tmp, monkeypatch, capsys):
    # neither the xlsx nor the CSV fallback can be written -> the approval
    # cannot be recorded -> it must not stamp
    monkeypatch.setenv("SEE_PROOF_LOG", str(in_tmp / "no_such_dir" / "proof_log.xlsx"))
    rc = _run(monkeypatch, canned_res(), "Jane Client")
    assert rc == 1
    assert "could not be logged" in capsys.readouterr().out
    assert not os.path.exists("F1_PROOF_vC1_APPROVED.html")


def test_approve_log_save_failure_refused(in_tmp, monkeypatch, capsys):
    def boom(*a, **k):
        raise OSError("disk full")
    monkeypatch.setattr(mp, "log_proof", boom)
    rc = _run(monkeypatch, canned_res(), "Jane Client")
    assert rc == 1
    out = capsys.readouterr().out
    assert "could not be logged" in out and "disk full" in out
    assert not os.path.exists("F1_PROOF_vC1_APPROVED.html")


def test_approve_with_ack_records_reason_in_html_and_log(in_tmp, monkeypatch, capsys):
    import openpyxl
    res = canned_res("REVIEW", {"size": ("PASS", "ok"),
                                "spelling": ("WARN", "1 word to review")})
    assert _run(monkeypatch, res, "Jane Client", ack="checked manually with the client") == 0
    doc = open("F1_PROOF_vC1_APPROVED.html", encoding="utf-8").read()
    assert "APPROVED" in doc and "Jane Client" in doc
    assert "checked manually with the client" in doc      # reason on the proof
    rows = list(openpyxl.load_workbook("proof_log.xlsx").active.iter_rows(values_only=True))
    status_col = rows[0].index("Status")
    assert any("REVIEW acknowledged: checked manually with the client" in str(r[status_col])
               for r in rows[1:])                          # reason in the log row
    assert "APPROVED by Jane Client" in capsys.readouterr().out


def test_clean_pass_approval_still_works(in_tmp, monkeypatch, capsys):
    import openpyxl
    assert _run(monkeypatch, canned_res(), "Jane Client") == 0
    assert os.path.exists("F1_PROOF_vC1_APPROVED.html")
    rows = list(openpyxl.load_workbook("proof_log.xlsx").active.iter_rows(values_only=True))
    assert rows[1:], "approval must be logged"


def test_non_approve_review_proof_generation_unchanged(in_tmp, monkeypatch, capsys):
    res = canned_res("REVIEW", {"size": ("PASS", "ok"), "spelling": ("WARN", "x")})
    assert _run(monkeypatch, res, None) == 0
    assert os.path.exists("F1_PROOF_vC1.html")
    assert "awaiting client sign-off" in capsys.readouterr().out
